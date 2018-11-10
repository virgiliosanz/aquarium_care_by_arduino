#/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from pprint import pprint as pp

wb = load_workbook(filename='NuevaPantalla.xlsx')
ws = wb['Design']
#print(ws['AH24'])
#print(ws['AH24'].value)


# From C17 to V17 - From C17 to C20 ==> 3,17 to 3,20 - 3,17 - 3,20,
# restar 13 a la row para sacar el "Color"
#
# Datos de color de primeras: AF3 to AF7 (nombre = hash de colores) AF=32
# Lúmenes en AH (34)
# Voltios en AL (35)
# Leds por Fase AC (row=15, column=29)
# UV en AF (32), IR en AG (33), W en AH (34), B en AI (35), R en AJ, (36) G en # AK (37)
# las fases empiezan en row 15

# Limpiamos resultados anteriores e iniciamos a 0
for row in range(15, 25):
    for col in range(29, 38):
        ws.cell(row=row, column=col, value=0)

colors = {}
for row in range(3, 9):
    color = ws.cell(row=row, column=32)
    lumens = ws.cell(row=row, column=34)
    volts = ws.cell(row=row, column=38)
    current = ws.cell(row=row, column=39)
    colors[color.value] = {'color': color.value,
                           'lumens': lumens.value,
                           'volts': volts.value,
                           'current': current.value}
# Columnas para el resultado
colors['UV']['col'] = 32
colors['IR']['col'] = 33
colors['W']['col'] = 34
colors['B']['col'] = 35
colors['R']['col'] = 36
colors['G']['col'] = 37

info = {}
for row in range(17, 21):
    for col in range(3, 23):
        fase = ws.cell(row=row, column=col)
        color = ws.cell(row=row - 13, column=col)

        fase_row = 14 + fase.value

#        print(row, col, fase.value, color.value, fase_row)

        total  = ws.cell(row=fase_row, column=29)
        ws.cell(row=fase_row, column=29, value=total.value + 1)

        volts  = ws.cell(row=fase_row, column=30)
        ws.cell(row=fase_row, column=30, value=volts.value + colors[color.value]['volts'])

        lumens  = ws.cell(row=fase_row, column=31)
        ws.cell(row=fase_row, column=31, value=lumens.value +
                colors[color.value]['lumens'])

        leds  = ws.cell(row=fase_row, column=colors[color.value]['col'])
        ws.cell(row=fase_row, column=colors[color.value]['col'],
                value=leds.value + 1)

wb.save("NuevaPantalla.xlsx")
